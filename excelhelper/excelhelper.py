import openpyxl
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog


class ExcelManipulator:

    def __init__(self, data_sheet_path, cryo_master_sheet_path, network_location):
        self.data_sheet_path = r"C:\Users\abehl\Downloads\Data Sheet A.xlsx"
        self.cryo_master_sheet_path = r"C:\Users\abehl\Downloads\CryoMaster.xlsx"
        self.network_location = r"C:\Users\abehl\Downloads"
        self.data_wb = openpyxl.load_workbook(self.data_sheet_path)
        self.cryo_master_wb = openpyxl.load_workbook(
            self.cryo_master_sheet_path)
    """
    def get_capsule_id(self, slide_position):
        ws = self.data_wb.active
        # Assuming column B has the slide position numbers
        for row in ws.iter_rows(min_col=2, max_col=3):
            if row[0].value == slide_position:
                capsule_id = row[1].value
                return row[1].value, capsule_id
    """

    def get_initial_values(self):
        while True:
            try:
                slide_position = int(
                    input("Please enter a slide position between 1 and 50\n> "))
                break
            except:
                if slide_position not in range(1, 51):
                    print("Invalid input, please try again")
                    break

        while True:
            try:
                pin_number = int(
                    input("Please enter a 4 digit pin number\n> "))
                break
            except ValueError:
                if not isinstance(pin_number, int) or not (1000 <= pin_number <= 9999):
                    raise ValueError("Pin number must be a 4 digit integer")
                else:
                    print("Pin number must be a 4 digit integer")
        return slide_position, pin_number

    def get_sheet(user_input: str) -> pd.DataFrame:
        sheet_name = "Final Data"
        file_path = filedialog.askopenfilename()
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        # Save the results to variables before returning
        csv_result = pd.read_csv(file_path)
        excel_result = pd.read_excel(file_path, sheet_name=sheet_name)
        return df, csv_result, excel_result

    def get_capsule_id(self, slide_position) -> str:
        # Takes thecryo_master_sheet_pathh
        print(self.data_sheet_path)
        print(self.network_location)
        print(self.cryo_master_sheet_path)
        data = pd.read_excel(io=self.cryo_master_sheet_path, header=9,
                             usecols="C:D", skiprows=14)
        # 0-based index, so 2 and 3 correspond to "C" and "D"
        c_column, d_column = data.columns[2], data.columns[3]
        for idx, entry in enumerate(data[c_column]):
            try:
                last_digits = int(str(entry).split('-')[-1])
            except ValueError:
                continue

            if last_digits == slide_position:
                return data[d_column].iloc[idx]
            return "No match found"

    def get_dimensional_measurements(self, slide_position):
        ws = self.data_wb.active
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=8, max_col=9):
            if col[0].value == slide_position:
                # Return OD and Wall thickness
                od = col[9].value
                wall_thickness = col[8].value
                return od, wall_thickness

    def add_data_to_cryo_master(self, slide_position):
        capsule_id = self.get_capsule_id(slide_position)
        od, wall_thickness = self.get_dimensional_measurements(slide_position)
        ws = self.cryo_master_wb.active
        ws.append([capsule_id, od, wall_thickness])
        self.cryo_master_wb.save(self.cryo_master_sheet_path)

    def create_folder_and_notes(self, capsule_id, pin_number):
        folder_name = f"{capsule_id} {pin_number} 1E"
        full_path = f"{self.network_location}/{folder_name}"
        os.makedirs(full_path, exist_ok=True)
        notes_content = f"{capsule_id} {pin_number} 1E\n" \
                        "0 deg  defects between 5um and 10um\n" \
                        "90 deg  defects between 5um and 10um\n" \
                        "180 deg  defects between 5um and 10um\n" \
                        "270 deg  defects between 5um and 10um"
        with open(f"{full_path}/notes.txt", "w") as f:
            f.write(notes_content)
